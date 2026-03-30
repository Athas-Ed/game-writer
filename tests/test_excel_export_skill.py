import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "skills/excel_export/scripts"


@pytest.fixture
def ex_mod():
    p = str(SCRIPTS)
    if p not in sys.path:
        sys.path.insert(0, p)
    import md_extract as me  # noqa
    import excel_writer as ew  # noqa

    return me, ew


def test_extract_dialogue_rows(ex_mod):
    me, _ = ex_mod
    text = "林夕：你好\n注意：这是说明\n艾莉：嗯"
    rows = me.extract_dialogue_rows(text)
    assert len(rows) == 2
    assert rows[0][0] == "林夕" and "你好" in rows[0][1]


def test_extract_first_gfm_table(ex_mod):
    me, _ = ex_mod
    t = "| 角色 | 台词 |\n| --- | --- |\n| A | 一 |\n| B | 二 |\n"
    h, r = me.extract_first_gfm_table(t)
    assert h == ["角色", "台词"]
    assert len(r) == 2


def test_extract_task_sections(ex_mod):
    me, _ = ex_mod
    text = """## 寻信
目标：找到信
奖励：金币

### 支线
目标：聊天
"""
    tasks = me.extract_task_rows(text)
    assert len(tasks) >= 1
    assert any("寻信" in t["name"] or "支线" in t["name"] for t in tasks)


def test_write_dialogue_xlsx(tmp_path, ex_mod):
    _, ew = ex_mod
    p = tmp_path / "a.xlsx"
    ew.write_dialogue_xlsx(p, [("X", "line")])
    from openpyxl import load_workbook

    wb = load_workbook(p)
    ws = wb.active
    assert ws["B2"].value == "X"
